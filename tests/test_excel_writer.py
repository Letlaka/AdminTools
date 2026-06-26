from __future__ import annotations

from datetime import date
import sys
from pathlib import Path
import tempfile
import unittest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from lib.department_mapping import DepartmentMatcher
from lib.report_model import build_report_bundle

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


@unittest.skipIf(load_workbook is None, "openpyxl is not installed")
class ExcelWriterTests(unittest.TestCase):
    def _sheet_values(self, worksheet) -> list[object]:
        values: list[object] = []
        for row in worksheet.iter_rows(values_only=True):
            values.extend(row)
        return values

    def test_creates_consolidated_workbook(self) -> None:
        from lib.excel_writer import create_consolidated_workbook

        records = [
            {
                "ComputerType": "Server",
                "CN": "SRV01",
                "DNSHostName": "srv01.example.local",
                "OUPath": "example/Health",
                "DistinguishedName": "CN=SRV01,OU=Health,DC=example,DC=local",
                "OperatingSystem": "Windows Server 2008 R2",
                "Description": "Primary app server",
                "LastLogonDate": "2026-06-20T14:30:00",
            }
        ]
        matcher = DepartmentMatcher(departments=["Health"], code_map={})
        bundle = build_report_bundle(records=records, matcher=matcher, departments=["Health"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "dashboard.xlsx"
            create_consolidated_workbook(
                output_path=path,
                bundle=bundle,
                financial_year="FY2026-2027",
                run_date_label=date(2026, 6, 25).isoformat(),
            )
            workbook = load_workbook(path)
            self.assertIn("Dashboard", workbook.sheetnames)
            self.assertIn("All Servers", workbook.sheetnames)
            dashboard = workbook["Dashboard"]
            dashboard_values = self._sheet_values(dashboard)
            self.assertIn("Server OS Lifecycle Summary", dashboard_values)
            self.assertIn("Legacy Server Spotlight", dashboard_values)
            self.assertIn("OSLifecycleBucket", dashboard_values)
            all_servers = workbook["All Servers"]
            headers = [cell.value for cell in all_servers[3]]
            self.assertIn("OSFamily", headers)
            self.assertIn("InactivityDays", headers)
            self.assertIn("InactivityStatus", headers)
            self.assertNotIn("StaleDays", headers)
            self.assertLess(headers.index("InactivityDays"), headers.index("LastSeenDate"))
            self.assertLess(headers.index("InactivityStatus"), headers.index("DaysSinceLastSeen"))
            last_seen_value = all_servers.cell(row=4, column=headers.index("LastSeenDate") + 1).value
            self.assertEqual(last_seen_value.date() if hasattr(last_seen_value, "date") else last_seen_value, date(2026, 6, 20))

    def test_dashboard_places_charts_next_to_source_tables(self) -> None:
        from lib.excel_writer import create_consolidated_workbook

        records = [
            {
                "ComputerType": "Server",
                "CN": "SRV2008",
                "DNSHostName": "srv2008.example.local",
                "OUPath": "example/Health",
                "DistinguishedName": "CN=SRV2008,OU=Health,DC=example,DC=local",
                "OperatingSystem": "Windows Server 2008 R2",
            },
            {
                "ComputerType": "Workstation",
                "CN": "PCWIN10",
                "DNSHostName": "pcwin10.example.local",
                "OUPath": "example/Health",
                "DistinguishedName": "CN=PCWIN10,OU=Health,DC=example,DC=local",
                "OperatingSystem": "Windows 10 Enterprise",
                "OperatingSystemVersion": "10.0.19045",
            },
        ]
        matcher = DepartmentMatcher(departments=["Health"], code_map={})
        bundle = build_report_bundle(records=records, matcher=matcher, departments=["Health"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "dashboard.xlsx"
            create_consolidated_workbook(
                output_path=path,
                bundle=bundle,
                financial_year="FY2026-2027",
                run_date_label=date(2026, 6, 25).isoformat(),
            )
            workbook = load_workbook(path)
            dashboard = workbook["Dashboard"]
            title_rows = {
                cell.value: cell.row
                for row in dashboard.iter_rows(min_col=1, max_col=1)
                for cell in row
                if cell.value
            }

            expected_titles = [
                "Department Device Summary",
                "Legacy Device Summary By Department",
                "Legacy Server Spotlight",
                "Server OS Lifecycle Summary",
                "Workstation OS Lifecycle Summary",
            ]
            self.assertEqual(len(dashboard._charts), len(expected_titles))
            self.assertEqual(
                [(chart.anchor._from.col, chart.anchor._from.row) for chart in dashboard._charts],
                [
                    (7, title_rows["Department Device Summary"] - 1),
                    (5, title_rows["Legacy Device Summary By Department"] - 1),
                    (3, title_rows["Legacy Server Spotlight"] - 1),
                    (4, title_rows["Server OS Lifecycle Summary"] - 1),
                    (4, title_rows["Workstation OS Lifecycle Summary"] - 1),
                ],
            )
            for chart, next_title in zip(dashboard._charts, expected_titles[1:]):
                self.assertGreaterEqual(title_rows[next_title], chart.anchor._from.row + 18)

    def test_creates_department_workbook_with_os_summary(self) -> None:
        from lib.excel_writer import create_department_workbook

        records = [
            {
                "ComputerType": "Server",
                "CN": "SRV01",
                "DNSHostName": "srv01.example.local",
                "OUPath": "example/Health",
                "DistinguishedName": "CN=SRV01,OU=Health,DC=example,DC=local",
                "OperatingSystem": "Windows Server 2012 R2",
                "LastLogonDate": "2026-06-15",
            },
            {
                "ComputerType": "Workstation",
                "CN": "PC01",
                "DNSHostName": "pc01.example.local",
                "OUPath": "example/Health",
                "DistinguishedName": "CN=PC01,OU=Health,DC=example,DC=local",
                "OperatingSystem": "Windows 11 Enterprise",
                "OperatingSystemVersion": "10.0.22631",
                "LastLogonDate": "2026-06-01",
            },
        ]
        matcher = DepartmentMatcher(departments=["Health"], code_map={})
        bundle = build_report_bundle(records=records, matcher=matcher, departments=["Health"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "department.xlsx"
            create_department_workbook(
                output_path=path,
                department="Health",
                server_records=bundle.records_for_type("Server"),
                workstation_records=bundle.records_for_type("Workstation"),
                financial_year="FY2026-2027",
                run_date_label=date(2026, 6, 25).isoformat(),
                department_summary=bundle.department_summary("Health"),
            )
            workbook = load_workbook(path)
            self.assertIn("Summary", workbook.sheetnames)
            summary = workbook["Summary"]
            summary_values = self._sheet_values(summary)
            self.assertIn("Server OS Lifecycle Summary", summary_values)
            self.assertIn("Workstation OS Lifecycle Summary", summary_values)
            self.assertIn("Legacy Server Spotlight", summary_values)
            servers_sheet = workbook["Servers"]
            server_headers = [cell.value for cell in servers_sheet[3]]
            server_last_seen = servers_sheet.cell(row=4, column=server_headers.index("LastSeenDate") + 1).value
            self.assertEqual(server_last_seen.date() if hasattr(server_last_seen, "date") else server_last_seen, date(2026, 6, 15))
            workstations_sheet = workbook["Workstations"]
            workstation_headers = [cell.value for cell in workstations_sheet[3]]
            workstation_last_seen = workstations_sheet.cell(row=4, column=workstation_headers.index("LastSeenDate") + 1).value
            self.assertEqual(workstation_last_seen.date() if hasattr(workstation_last_seen, "date") else workstation_last_seen, date(2026, 6, 1))


if __name__ == "__main__":
    unittest.main()
