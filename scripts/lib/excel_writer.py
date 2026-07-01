from __future__ import annotations

from datetime import datetime
import math
import re
from pathlib import Path

from lib.report_model import LEGACY_SERVER_SPOTLIGHT_BUCKETS, serialize_for_sheet
from lib.spreadsheet_safety import safe_cell_value

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise RuntimeError(
        "openpyxl is required for Excel report generation. "
        "Install it with `uv sync` or `python -m pip install -r requirements-reporting.txt`."
    ) from exc


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
LEGACY_FILL = PatternFill("solid", fgColor="FDE68A")
SPOTLIGHT_FILL = PatternFill("solid", fgColor="FECACA")
DASHBOARD_SECTION_GAP_ROWS = 3
DASHBOARD_CHART_GAP_COLUMNS = 1
DASHBOARD_CHART_ROW_FACTOR = 2


def sanitize_file_component(value: str) -> str:
    sanitized = re.sub(r'[<>:"/\\\\|?*]+', "_", value).strip()
    sanitized = re.sub(r"\s+", " ", sanitized)
    return sanitized or "Unknown"


def safe_sheet_name(value: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\*:/\\?]+", "_", value).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 1
    while candidate in existing:
        suffix_text = f"_{suffix}"
        candidate = f"{cleaned[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    existing.add(candidate)
    return candidate


def _format_value(value: object) -> object:
    if isinstance(value, datetime):
        return value
    return safe_cell_value(value)


def _write_cell(worksheet, row: int, column: int, value: object):
    formatted_value = _format_value(value)
    cell = worksheet.cell(row=row, column=column, value=formatted_value)
    if isinstance(formatted_value, str):
        cell.data_type = "s"
    return cell


def _autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _write_rows_table(
    worksheet,
    rows: list[dict[str, object]],
    start_row: int,
    title: str,
    highlight_legacy_column: str | None = None,
    spotlight_column: str | None = None,
) -> tuple[int, int, list[str]]:
    _write_cell(worksheet, row=start_row, column=1, value=title)
    worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=14)

    headers = list(rows[0].keys()) if rows else []
    header_row = start_row + 2

    if not headers:
        _write_cell(worksheet, row=header_row, column=1, value="No records")
        return header_row, header_row, []

    for column_index, header in enumerate(headers, start=1):
        cell = _write_cell(worksheet, row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    last_data_row = header_row
    for row_offset, row in enumerate(rows, start=1):
        current_row = header_row + row_offset
        last_data_row = current_row
        is_legacy = (
            highlight_legacy_column is not None
            and str(row.get(highlight_legacy_column) or "").strip().lower() in {"yes", "true"}
        )
        is_spotlight = (
            spotlight_column is not None
            and str(row.get(spotlight_column) or "") in LEGACY_SERVER_SPOTLIGHT_BUCKETS
        )
        fill = SPOTLIGHT_FILL if is_spotlight else LEGACY_FILL if is_legacy else None

        for column_index, header in enumerate(headers, start=1):
            cell = _write_cell(worksheet, row=current_row, column=column_index, value=row.get(header))
            if fill is not None:
                cell.fill = fill

    return header_row, last_data_row, headers


def _dashboard_chart_anchor(headers: list[str], start_row: int) -> str:
    anchor_column = max(len(headers) + DASHBOARD_CHART_GAP_COLUMNS + 1, 4)
    return f"{get_column_letter(anchor_column)}{start_row}"


def _next_dashboard_section_row(table_last_row: int, chart_start_row: int, chart_height: float) -> int:
    chart_last_row = chart_start_row + math.ceil(chart_height * DASHBOARD_CHART_ROW_FACTOR)
    return max(table_last_row, chart_last_row) + DASHBOARD_SECTION_GAP_ROWS


def _add_bar_chart(
    worksheet,
    title: str,
    y_axis_title: str,
    x_axis_title: str,
    header_row: int,
    last_row: int,
    data_column: int,
    category_column: int,
    anchor_cell: str,
    height: float = 7.5,
    width: float = 14,
) -> None:
    if last_row <= header_row:
        return

    chart = BarChart()
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    data = Reference(worksheet, min_col=data_column, min_row=header_row, max_row=last_row)
    categories = Reference(worksheet, min_col=category_column, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = height
    chart.width = width
    worksheet.add_chart(chart, anchor_cell)


def _write_table(worksheet, rows: list[dict[str, object]], title: str | None = None) -> None:
    current_row = 1
    if title:
        _write_cell(worksheet, row=current_row, column=1, value=title)
        worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 2

    headers = list(rows[0].keys()) if rows else []
    if not headers:
        _write_cell(worksheet, row=current_row, column=1, value="No records")
        return

    header_row = current_row
    data_start_row = current_row + 1

    for column_index, header in enumerate(headers, start=1):
        cell = _write_cell(worksheet, row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    os_legacy_index = headers.index("OSIsLegacy") + 1 if "OSIsLegacy" in headers else None
    for row_index, row in enumerate(rows, start=data_start_row):
        row_is_legacy = bool(row.get("OSIsLegacy"))
        for column_index, header in enumerate(headers, start=1):
            cell = _write_cell(worksheet, row=row_index, column=column_index, value=row.get(header))
            if row_is_legacy:
                cell.fill = LEGACY_FILL

        if os_legacy_index is not None:
            _write_cell(worksheet, row=row_index, column=os_legacy_index, value="Yes" if row_is_legacy else "No")

    worksheet.freeze_panes = worksheet.cell(row=data_start_row, column=1)
    worksheet.auto_filter.ref = worksheet.dimensions
    _autofit_columns(worksheet)


def create_consolidated_workbook(
    output_path: Path,
    bundle,
    financial_year: str,
    run_date_label: str,
) -> None:
    workbook = Workbook()
    sheet_name_registry: set[str] = set()

    dashboard = workbook.active
    dashboard.title = "Dashboard"
    sheet_name_registry.add("Dashboard")
    _write_dashboard(dashboard, bundle, financial_year, run_date_label)

    all_servers = workbook.create_sheet(title=safe_sheet_name("All Servers", sheet_name_registry))
    _write_table(all_servers, serialize_for_sheet(bundle.records_for_type("Server")), title="All Servers")

    all_workstations = workbook.create_sheet(title=safe_sheet_name("All Workstations", sheet_name_registry))
    _write_table(
        all_workstations,
        serialize_for_sheet(bundle.records_for_type("Workstation")),
        title="All Workstations",
    )

    for department in bundle.departments_with_unknown:
        worksheet = workbook.create_sheet(title=safe_sheet_name(department, sheet_name_registry))
        _write_table(
            worksheet,
            serialize_for_sheet(bundle.records_for_department(department)),
            title=f"{department} Devices",
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_dashboard(worksheet, bundle, financial_year: str, run_date_label: str) -> None:
    worksheet["A1"] = "Active Directory Device Dashboard"
    worksheet["A1"].font = Font(bold=True, size=16)
    worksheet["A2"] = "Financial Year"
    worksheet["B2"] = financial_year
    worksheet["A3"] = "Run Date"
    worksheet["B3"] = run_date_label
    worksheet["A4"] = "Total Servers"
    worksheet["B4"] = len(bundle.records_for_type("Server"))
    worksheet["A5"] = "Total Workstations"
    worksheet["B5"] = len(bundle.records_for_type("Workstation"))
    worksheet["A6"] = "Total Devices"
    worksheet["B6"] = len(bundle.records)

    department_start_row = 8
    department_rows = bundle.dashboard_rows()
    department_header_row, department_last_row, department_headers = _write_rows_table(
        worksheet,
        department_rows,
        start_row=department_start_row,
        title="Department Device Summary",
    )
    _add_bar_chart(
        worksheet,
        title="Devices Per Department",
        y_axis_title="Devices",
        x_axis_title="Department",
        header_row=department_header_row,
        last_row=department_last_row,
        data_column=4,
        category_column=1,
        anchor_cell=_dashboard_chart_anchor(department_headers, department_start_row),
        width=16,
    )

    legacy_start_row = _next_dashboard_section_row(department_last_row, department_start_row, 7.5)
    legacy_rows = bundle.legacy_department_rows()
    legacy_header_row, legacy_last_row, legacy_headers = _write_rows_table(
        worksheet,
        legacy_rows,
        start_row=legacy_start_row,
        title="Legacy Device Summary By Department",
    )
    _add_bar_chart(
        worksheet,
        title="Legacy Devices By Department",
        y_axis_title="Legacy Devices",
        x_axis_title="Department",
        header_row=legacy_header_row,
        last_row=legacy_last_row,
        data_column=4,
        category_column=1,
        anchor_cell=_dashboard_chart_anchor(legacy_headers, legacy_start_row),
    )

    spotlight_start_row = _next_dashboard_section_row(legacy_last_row, legacy_start_row, 7.5)
    spotlight_rows = bundle.legacy_spotlight_rows()
    spotlight_header_row, spotlight_last_row, spotlight_headers = _write_rows_table(
        worksheet,
        spotlight_rows,
        start_row=spotlight_start_row,
        title="Legacy Server Spotlight",
        spotlight_column="OSLifecycleBucket",
    )
    _add_bar_chart(
        worksheet,
        title="Legacy Server Spotlight",
        y_axis_title="Servers",
        x_axis_title="OS Bucket",
        header_row=spotlight_header_row,
        last_row=spotlight_last_row,
        data_column=2,
        category_column=1,
        anchor_cell=_dashboard_chart_anchor(spotlight_headers, spotlight_start_row),
    )

    server_start_row = _next_dashboard_section_row(spotlight_last_row, spotlight_start_row, 7.5)
    server_os_rows = bundle.os_summary_rows("Server")
    server_header_row, server_last_row, server_headers = _write_rows_table(
        worksheet,
        server_os_rows,
        start_row=server_start_row,
        title="Server OS Lifecycle Summary",
        highlight_legacy_column="Legacy",
        spotlight_column="OSLifecycleBucket",
    )
    _add_bar_chart(
        worksheet,
        title="Server Lifecycle Distribution",
        y_axis_title="Servers",
        x_axis_title="OS Bucket",
        header_row=server_header_row,
        last_row=server_last_row,
        data_column=2,
        category_column=1,
        anchor_cell=_dashboard_chart_anchor(server_headers, server_start_row),
        height=8.5,
    )

    workstation_start_row = _next_dashboard_section_row(server_last_row, server_start_row, 8.5)
    workstation_os_rows = bundle.os_summary_rows("Workstation")
    workstation_header_row, workstation_last_row, workstation_headers = _write_rows_table(
        worksheet,
        workstation_os_rows,
        start_row=workstation_start_row,
        title="Workstation OS Lifecycle Summary",
        highlight_legacy_column="Legacy",
    )
    _add_bar_chart(
        worksheet,
        title="Workstation Lifecycle Distribution",
        y_axis_title="Workstations",
        x_axis_title="OS Bucket",
        header_row=workstation_header_row,
        last_row=workstation_last_row,
        data_column=2,
        category_column=1,
        anchor_cell=_dashboard_chart_anchor(workstation_headers, workstation_start_row),
        height=8,
    )

    worksheet.freeze_panes = "A8"
    _autofit_columns(worksheet)


def create_department_workbook(
    output_path: Path,
    department: str,
    server_records: list[dict[str, object]],
    workstation_records: list[dict[str, object]],
    financial_year: str,
    run_date_label: str,
    department_summary: dict[str, object],
) -> None:
    workbook = Workbook()
    registry: set[str] = set()

    summary = workbook.active
    summary.title = "Summary"
    registry.add("Summary")
    summary["A1"] = f"{department} Device Report"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A2"] = "Financial Year"
    summary["B2"] = financial_year
    summary["A3"] = "Run Date"
    summary["B3"] = run_date_label
    summary["A4"] = "Servers"
    summary["B4"] = department_summary["Servers"]
    summary["A5"] = "Workstations"
    summary["B5"] = department_summary["Workstations"]
    summary["A6"] = "Total"
    summary["B6"] = department_summary["Total"]
    summary["A7"] = "Legacy Servers"
    summary["B7"] = department_summary["LegacyServers"]
    summary["A8"] = "Legacy Workstations"
    summary["B8"] = department_summary["LegacyWorkstations"]

    _, server_last_row, _ = _write_rows_table(
        summary,
        department_summary["ServerOSSummary"],
        start_row=10,
        title="Server OS Lifecycle Summary",
        highlight_legacy_column="Legacy",
        spotlight_column="OSLifecycleBucket",
    )

    _, workstation_last_row, _ = _write_rows_table(
        summary,
        department_summary["WorkstationOSSummary"],
        start_row=server_last_row + 3,
        title="Workstation OS Lifecycle Summary",
        highlight_legacy_column="Legacy",
    )

    _write_rows_table(
        summary,
        department_summary["LegacySpotlight"],
        start_row=workstation_last_row + 3,
        title="Legacy Server Spotlight",
        spotlight_column="OSLifecycleBucket",
    )
    _autofit_columns(summary)

    servers_sheet = workbook.create_sheet(title=safe_sheet_name("Servers", registry))
    _write_table(servers_sheet, serialize_for_sheet(server_records), title=f"{department} Servers")

    workstations_sheet = workbook.create_sheet(title=safe_sheet_name("Workstations", registry))
    _write_table(
        workstations_sheet,
        serialize_for_sheet(workstation_records),
        title=f"{department} Workstations",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
